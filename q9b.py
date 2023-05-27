import openpyxl # pip install openpyxl

# Read data from a spreadsheet
def read_spreadsheet(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    return data

# Write data into a spreadsheet
def write_spreadsheet(file_name, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row in data:
        sheet.append(row)

    workbook.save(file_name)
    print("Data written to the spreadsheet successfully.")

# Test the program
file_name = "examples.xlsx"

# Read data from the spreadsheet
read_data = read_spreadsheet(file_name)
print("Data read from the spreadsheet:")
for row in read_data:
    print(row)

# Modify the data
modified_data = []
for row in read_data:
    modified_row = [cell.upper() if isinstance(cell, str) else cell for cell in row]
    modified_data.append(modified_row)

# Write modified data into a new spreadsheet
new_file_name = "modified_example.xlsx"
write_spreadsheet(new_file_name, modified_data)
